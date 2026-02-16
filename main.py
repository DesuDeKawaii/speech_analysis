from megafon import sync_calls_from_megafon
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from database import init_db, SessionLocal, Call
from datetime import datetime, timedelta
import random
import time
from faker import Faker

fake = Faker("ru_RU")
OPERATORS = ["Смирнова Анна", "Кузнецова Елена", "Васильева Мария"]

def create_mock_data():
    """Наполняет базу фейковыми звонками для теста"""
    session = SessionLocal()
    
    if session.query(Call).count() > 0:
        print("ℹ️ В базе уже есть данные. Используем их.")
        session.close()
        return

    print("🎲 Генерирую 30 тестовых звонков...")
    
    for _ in range(30):
        ai_mock = {
            "greeting": random.randint(4, 10),
            "needs": random.randint(3, 9),
            "presentation": random.randint(4, 10),
            "objection": random.randint(5, 10),
            "services_count": random.choice([0, 1, 1, 2]),
            "bonus": random.choice([0, 0, 0, 500]),
            "summary": fake.sentence(nb_words=15), # Чуть длиннее текст
            "recommendation": random.choice([
                "Не перебивать клиента, выслушать до конца.",
                "Предлагать доп. услуги (УЗИ, анализы) более настойчиво.",
                "Говорить громче и увереннее, клиент переспрашивает.",
                "Выучить прайс-лист, долгие паузы при поиске цены.",
                "Отличная работа, эталонный диалог."
            ])
        }

        call = Call(
            id=str(random.randint(100000, 999999)),
            date=datetime.now() - timedelta(days=random.randint(0, 14)),
            operator=random.choice(OPERATORS),
            duration=random.randint(60, 400),
            status="PROCESSED",
            ai_data=ai_mock
        )
        session.add(call)
    
    session.commit()
    print("✅ Тестовые данные сохранены в calls.db")
    session.close()

def apply_beautiful_styles(ws, start_row, end_row, last_col):
    """Наводит красоту: рамки, отступы, переносы строк"""
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Проходим по всем ячейкам с данными
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=last_col):
        for cell in row:
            # 1. Делаем рамку
            cell.border = thin_border
            
            # 2. Настраиваем выравнивание
            # Если колонка с длинным текстом (обычно это последние 2) - делаем перенос
            if cell.column >= last_col - 1: 
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            else:
                # Остальные по центру
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                
            # Имя оператора (1 колонка) - выравниваем по левому краю
            if cell.column == 1:
                cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)

def set_column_widths(ws, columns_config):
    """Задает ширину столбцов"""
    for col_letter, width in columns_config.items():
        ws.column_dimensions[col_letter].width = width

def generate_excel():
    print("📊 Формирую красивый Excel отчет...")
    session = SessionLocal()
    calls = session.query(Call).filter(Call.status == "PROCESSED").all()
    
    try:
        wb = load_workbook("template.xlsx")
    except FileNotFoundError:
        print("❌ ОШИБКА: Файл template.xlsx не найден!")
        return

    # ==========================================
    # ЛИСТ 1: Детальный отчет
    # ==========================================
    try:
        ws_detail = wb["Детальный отчет"]
    except KeyError:
        print("❌ ОШИБКА: Лист 'Детальный отчет' не найден")
        return

    start_row = 2
    data_for_pandas = []

    for i, call in enumerate(calls):
        r = start_row + i
        ai = call.ai_data
        
        mins, secs = divmod(call.duration, 60)
        duration_str = f"{mins}:{secs:02d}"
        
        ws_detail.cell(row=r, column=1, value=call.operator)
        ws_detail.cell(row=r, column=2, value=call.date.strftime("%d.%m.%Y %H:%M"))
        ws_detail.cell(row=r, column=3, value=duration_str)
        ws_detail.cell(row=r, column=4, value=ai.get('greeting', 0))
        ws_detail.cell(row=r, column=5, value=ai.get('needs', 0))
        ws_detail.cell(row=r, column=6, value=ai.get('presentation', 0))
        ws_detail.cell(row=r, column=7, value=ai.get('objection', 0))
        ws_detail.cell(row=r, column=8, value=ai.get('services_count', 0))
        ws_detail.cell(row=r, column=9, value=ai.get('bonus', 0))
        ws_detail.cell(row=r, column=10, value=ai.get('summary', '-'))
        ws_detail.cell(row=r, column=11, value=ai.get('recommendation', '-'))

        data_for_pandas.append({
            "Оператор": call.operator,
            "greeting": ai.get('greeting', 0),
            "needs": ai.get('needs', 0),
            "presentation": ai.get('presentation', 0),
            "objection": ai.get('objection', 0),
            "services": ai.get('services_count', 0),
            "recommendation": ai.get('recommendation', '-')
        })

    set_column_widths(ws_detail, {
        'A': 25, # Оператор
        'B': 18, # Дата
        'C': 10, # Длительность
        'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 10, 'I': 10, # Оценки
        'J': 50, # Анализ (широкий)
        'K': 50  # Рекомендации (широкий)
    })
    apply_beautiful_styles(ws_detail, start_row, start_row + len(calls) - 1, 11)


    # ==========================================
    # ЛИСТ 2: Общий отчет
    # ==========================================
    try:
        ws_summary = wb["Общий отчет"]
        if data_for_pandas:
            df = pd.DataFrame(data_for_pandas)
            
            # --- 1. Шапка ---
            ws_summary.cell(row=2, column=1, value=len(df))
            ws_summary.cell(row=2, column=2, value=df['services'].sum())
            avg_total = (df['greeting'].mean() + df['needs'].mean() + df['presentation'].mean() + df['objection'].mean()) / 4
            ws_summary.cell(row=2, column=3, value=round(avg_total, 2))
            
            # Стили для шапки (выравнивание по центру)
            for col in range(1, 4):
                ws_summary.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')

            # --- 2. Таблица операторов ---
            grouped = df.groupby("Оператор")
            start_row_sum = 16 
            
            current_row = start_row_sum
            for name, group in grouped:
                avg_kpi = (group['greeting'].mean() + group['needs'].mean() + 
                           group['presentation'].mean() + group['objection'].mean()) / 4
                
                status_text = "Золотой" if avg_kpi > 8.5 else "Серебряный" if avg_kpi >= 7 else "Медный"
                status_val = f"{avg_kpi:.2f}\n{status_text}"
                
                rec_mode = group['recommendation'].mode()
                top_rec = rec_mode[0] if not rec_mode.empty else "Нет данных"
                final_ai = f"Статус: {status_text}.\nЧастая ошибка: {top_rec}"

                ws_summary.cell(row=current_row, column=1, value=name)
                ws_summary.cell(row=current_row, column=2, value=len(group))
                ws_summary.cell(row=current_row, column=3, value=group['services'].sum())
                ws_summary.cell(row=current_row, column=4, value=status_val)
                ws_summary.cell(row=current_row, column=5, value=final_ai)

                current_row += 1
            
            # НАВОДИМ КРАСОТУ (Сводный)
            set_column_widths(ws_summary, {
                'A': 25, # Оператор
                'B': 15, # Звонки
                'C': 15, # Услуги
                'D': 20, # Статус
                'E': 60  # Рекомендации
            })
            apply_beautiful_styles(ws_summary, start_row_sum, current_row - 1, 5)

    except KeyError:
        print("⚠️ Лист 'Общий отчет' не найден.")

    # ==========================================
    # СОХРАНЕНИЕ
    # ==========================================
    # Формат: Report_16.02.26.xlsx
    date_str = datetime.now().strftime("%d.%m.%y")
    
    output_filename = f"Report_{date_str}.xlsx"  #подправить, если в один день сделали два отчёта (не стирается старый)
    
    try:
        wb.save(output_filename)
        print(f"🚀 УСПЕХ! Файл создан: {output_filename}")
    except PermissionError:
        print(f"⛔ ОШИБКА: Закрой файл {output_filename} и попробуй снова.")

if __name__ == "__main__":
    init_db()

    # 1. СБОР РЕАЛЬНЫХ ДАННЫХ
    # (Если ключи в .env неверные, он просто напишет ошибку, но не упадет)
    sync_calls_from_megafon(days_back=14)

    # 2. ГЕНЕРАЦИЯ ОТЧЕТА
    # Сейчас он создаст пустой отчет, так как у новых звонков статус "NEW",
    # а отчет строится по "PROCESSED". 
    # Следующим этапом мы напишем ИИ-обработчик, который переведет их в PROCESSED.
    generate_excel()